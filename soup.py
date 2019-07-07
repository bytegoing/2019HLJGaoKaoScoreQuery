from bs4 import BeautifulSoup
import xlrd
import openpyxl

data = xlrd.open_workbook('D:/model.xlsx')
table = data.sheets()[0]
nrows = table.nrows

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Sheet1"

print("共有"+str(nrows-1)+"名同学信息")
rawList = []
sheet.cell(1,1,['准考证号','姓名','语文','数学','外语','综合','总分','听力'])
for i in range(nrows):
    soup = BeautifulSoup(open('D:/GaoKaoScoreHTML/'+table.col_values(2)[i+1]+'.html', encoding='UTF-8'), features='lxml')
    raw = soup.select_one('tr[bgcolor="#F5F5F5"]').get_text()
    rawList = raw.split()
    print(rawList)
    for j in range(8):
        sheet.cell(row=i+2, column=j+2, value=rawList[j])

workbook.save('D:/final.xlsx')
